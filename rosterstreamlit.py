import streamlit as st
import openpyxl
import random
from openpyxl.styles import PatternFill
from io import BytesIO
import pandas as pd

# -------------------------------------------------------------------------
# 1) Session State Başlatma
# -------------------------------------------------------------------------
def initialize_session_states():
    """Streamlit oturumu başladığında (veya sıfırlandığında) varsayılan değerleri atar."""
    if "participants" not in st.session_state:
        st.session_state.participants = []

    # Sektörler
    if "boards" not in st.session_state:
        st.session_state.boards = ["SWN", "SWS", "SWF", "SCF", "SEF", "SEN", "SEC", "SES", "SAG"]

    # Günler
    if "days_of_week" not in st.session_state:
        st.session_state.days_of_week = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]

    # Zaman aralıkları
    if "timeslots" not in st.session_state:
        st.session_state.timeslots = ["09:00-10:00", "10:30-11:30", "13:00-14:00", "14:30-15:30"]

    # Standart senaryolar
    if "standard_scenarios" not in st.session_state:
        st.session_state.standard_scenarios = [
            "Kuzey Doğu Peak",
            "Kuzey Batı Peak",
            "Güney Doğu Peak",
            "Güney Batı Peak",
            "Ters Kuzey",
            "Ters Güney",
        ]

    # Asıl senaryo listesi: [ (senaryo_adı, gün, timeslot, tekrar), ... ]
    if "scenarios" not in st.session_state:
        st.session_state.scenarios = [
            ["Kuzey Doğu Peak", "Pazartesi", "09:00-10:00", 1],
            ["Kuzey Batı Peak", "Pazartesi", "10:30-11:30", 1],
            ["Güney Doğu Peak", "Pazartesi", "13:00-14:00", 1],
            ["Güney Batı Peak", "Pazartesi", "14:30-15:30", 1],
        ]

    # Mevcut atama yöntemleri
    if "assignment_methods" not in st.session_state:
        st.session_state.assignment_methods = [
            "Random",
            "Round Robin",
            "Balanced",
            "Constraint (Latin Square)"
        ]

    # Varsayılan seçili atama yöntemi
    if "selected_method" not in st.session_state:
        st.session_state.selected_method = "Round Robin"

    # Plan verileri
    if "plan_data" not in st.session_state:
        st.session_state.plan_data = []

# -------------------------------------------------------------------------
# 2) Katılımcılar
# -------------------------------------------------------------------------
def add_participant():
    """Yeni bir katılımcı eklerken en küçük kullanılmayan ATC indeksini bulur."""
    used_indices = sorted([
        int(p.replace("ATC", "")) for p in st.session_state.participants
        if p.startswith("ATC") and p[3:].isdigit()
    ])
    new_index = 1
    for idx in used_indices:
        if new_index < idx:
            break
        new_index = idx + 1
    new_name = f"ATC{new_index}"
    st.session_state.participants.append(new_name)

def remove_participant(selected_participant):
    """Seçilen katılımcıyı listeden çıkarır."""
    if selected_participant in st.session_state.participants:
        st.session_state.participants.remove(selected_participant)
        st.success(f"Katılımcı '{selected_participant}' silindi.")

# -------------------------------------------------------------------------
# 3) Sektörler (Boards)
# -------------------------------------------------------------------------
def add_board(new_board_name):
    """Yeni bir sektör ekler."""
    if new_board_name and new_board_name not in st.session_state.boards:
        st.session_state.boards.append(new_board_name)
        st.success(f"Sektör '{new_board_name}' eklendi.")
    elif new_board_name in st.session_state.boards:
        st.warning("Bu sektör zaten mevcut.")

def edit_board(old_name, new_name):
    """Mevcut bir sektörün adını günceller."""
    if old_name in st.session_state.boards:
        if new_name and new_name not in st.session_state.boards:
            idx = st.session_state.boards.index(old_name)
            st.session_state.boards[idx] = new_name
            st.success(f"Sektör '{old_name}' -> '{new_name}' olarak güncellendi.")
        else:
            st.warning("Yeni sektör adı boş veya zaten mevcut.")
    else:
        st.warning(f"Sektör '{old_name}' listede yok.")

def remove_board(selected_board):
    """Seçilen sektörü listeden çıkarır."""
    if selected_board in st.session_state.boards:
        st.session_state.boards.remove(selected_board)
        st.success(f"Sektör '{selected_board}' silindi.")

# -------------------------------------------------------------------------
# 4) Günler
# -------------------------------------------------------------------------
def add_day(new_day_name):
    """Yeni bir gün ekler."""
    if new_day_name and new_day_name not in st.session_state.days_of_week:
        st.session_state.days_of_week.append(new_day_name)
        st.success(f"Gün '{new_day_name}' eklendi.")
    elif new_day_name in st.session_state.days_of_week:
        st.warning("Bu gün zaten mevcut.")

def rename_day(old_day, new_day):
    """Mevcut bir gün adını günceller."""
    if old_day in st.session_state.days_of_week:
        if new_day and new_day not in st.session_state.days_of_week:
            idx = st.session_state.days_of_week.index(old_day)
            st.session_state.days_of_week[idx] = new_day
            st.success(f"Gün '{old_day}' -> '{new_day}' olarak güncellendi.")
        else:
            st.warning("Yeni gün zaten mevcut veya geçersiz.")
    else:
        st.warning(f"'{old_day}' mevcut günler içinde bulunamadı.")

def remove_day(selected_day):
    """Seçilen günü listeden çıkarır."""
    if selected_day in st.session_state.days_of_week:
        st.session_state.days_of_week.remove(selected_day)
        st.success(f"Gün '{selected_day}' silindi.")

# -------------------------------------------------------------------------
# 5) Zaman Aralıkları (Timeslots)
# -------------------------------------------------------------------------
def add_timeslot(new_slot):
    """Yeni bir zaman aralığı ekler."""
    if new_slot and new_slot not in st.session_state.timeslots:
        st.session_state.timeslots.append(new_slot)
        st.success(f"Zaman aralığı '{new_slot}' eklendi.")
    elif new_slot in st.session_state.timeslots:
        st.warning("Bu zaman aralığı zaten mevcut.")

def edit_timeslot(old_slot, new_slot):
    """Mevcut bir zaman aralığını günceller."""
    if old_slot in st.session_state.timeslots:
        if new_slot and new_slot not in st.session_state.timeslots:
            idx = st.session_state.timeslots.index(old_slot)
            st.session_state.timeslots[idx] = new_slot
            st.success(f"Zaman aralığı '{old_slot}' -> '{new_slot}' olarak güncellendi.")
        else:
            st.warning("Yeni zaman aralığı boş veya zaten mevcut.")
    else:
        st.warning(f"Zaman aralığı '{old_slot}' listede yok.")

def remove_timeslot(selected_slot):
    """Seçilen zaman aralığını listeden çıkarır."""
    if selected_slot in st.session_state.timeslots:
        st.session_state.timeslots.remove(selected_slot)
        st.success(f"Zaman aralığı '{selected_slot}' silindi.")

# -------------------------------------------------------------------------
# 6) Standart Senaryolar
# -------------------------------------------------------------------------
def add_standard_scenario(new_scenario):
    """Yeni bir standart senaryo ekler."""
    if new_scenario and new_scenario not in st.session_state.standard_scenarios:
        st.session_state.standard_scenarios.append(new_scenario)
        st.success(f"Standart senaryo '{new_scenario}' eklendi.")
    elif new_scenario in st.session_state.standard_scenarios:
        st.warning("Bu standart senaryo zaten mevcut.")

def edit_standard_scenario(old_scenario, new_scenario):
    """Mevcut bir standart senaryonun adını günceller."""
    if old_scenario in st.session_state.standard_scenarios:
        if new_scenario and new_scenario not in st.session_state.standard_scenarios:
            idx = st.session_state.standard_scenarios.index(old_scenario)
            st.session_state.standard_scenarios[idx] = new_scenario
            st.success(f"Standart senaryo '{old_scenario}' -> '{new_scenario}' olarak güncellendi.")
        else:
            st.warning("Yeni standart senaryo adı boş veya zaten mevcut.")
    else:
        st.warning(f"Standart senaryo '{old_scenario}' listede yok.")

def remove_standard_scenario(selected_scenario):
    """Seçilen standart senaryoyu listeden çıkarır."""
    if selected_scenario in st.session_state.standard_scenarios:
        st.session_state.standard_scenarios.remove(selected_scenario)
        st.success(f"Standart senaryo '{selected_scenario}' silindi.")

# -------------------------------------------------------------------------
# 7) Asıl Senaryolar
# -------------------------------------------------------------------------
def add_scenario():
    """Yeni bir senaryo ekler (varsayılan ilk senaryo, ilk gün, ilk zaman)."""
    default_s = st.session_state.standard_scenarios[0] if st.session_state.standard_scenarios else "Senaryo?"
    default_d = st.session_state.days_of_week[0] if st.session_state.days_of_week else "Gün?"
    default_t = st.session_state.timeslots[0] if st.session_state.timeslots else "Zaman?"
    st.session_state.scenarios.append([default_s, default_d, default_t, 1])

def remove_scenario(index):
    """Belirli bir indeksteki senaryoyu siler."""
    if 0 <= index < len(st.session_state.scenarios):
        st.session_state.scenarios.pop(index)

# -------------------------------------------------------------------------
# 8) Atama Yöntemleri (Plan Oluşturma)
# -------------------------------------------------------------------------
def assign_random(scenario_list, participants, boards):
    plan_data = []
    p_copy = participants[:]
    for (day, slot, sc_name) in scenario_list:
        random.shuffle(p_copy)
        assignment = []
        for b_i in range(len(boards)):
            assignment.append(p_copy[b_i % len(p_copy)])
        plan_data.append((day, slot, sc_name, assignment))
    return plan_data

def assign_round_robin(scenario_list, participants, boards):
    plan_data = []
    p_count = len(participants)
    for s_idx, (day, slot, sc_name) in enumerate(scenario_list):
        assignment = []
        for b_i in range(len(boards)):
            idx = (b_i + s_idx) % p_count
            assignment.append(participants[idx])
        plan_data.append((day, slot, sc_name, assignment))
    return plan_data

def assign_balanced(scenario_list, participants, boards):
    plan_data = []
    participant_count = {p: 0 for p in participants}
    participant_board_count = {(p, b): 0 for p in participants for b in boards}

    for (day, slot, sc_name) in scenario_list:
        assignment = []
        for b in boards:
            best_p = min(participants, key=lambda p: (participant_count[p] + participant_board_count[(p,b)]))
            assignment.append(best_p)
            participant_count[best_p] += 1
            participant_board_count[(best_p, b)] += 1
        plan_data.append((day, slot, sc_name, assignment))
    return plan_data

def assign_constraint_latin(scenario_list, participants, boards):
    n_boards = len(boards)
    n_parts = len(participants)
    m_scenario = len(scenario_list)

    # Latin Square yöntemi için, katılımcı sayısı == board sayısı olmalı
    # ve senaryo sayısı en az board sayısı kadar olmalı
    if n_boards != n_parts:
        st.error("Hata: Constraint (Latin Square) için '#participants == #boards' olmalı.")
        return None
    if m_scenario < n_boards:
        st.error("Hata: Constraint (Latin Square) için 'toplam senaryo sayısı >= board sayısı' olmalı.")
        return None

    plan_data = []

    # İlk n_boards satır => Latin kare
    for row in range(n_boards):
        if row >= m_scenario:
            break
        (day, slot, sc_name) = scenario_list[row]
        assignment = []
        for col in range(n_boards):
            part_idx = (row + col) % n_boards
            assignment.append(participants[part_idx])
        plan_data.append((day, slot, sc_name, assignment))

    # Kalan senaryolar
    participant_usage = {p: n_boards for p in participants}
    for row in range(n_boards, m_scenario):
        (day, slot, sc_name) = scenario_list[row]
        available = [p for p in participants if participant_usage[p] < m_scenario]
        if len(available) < n_boards:
            st.error(f"Kalan senaryolarda yetersiz katılımcı kaldı (satır={row}).")
            return None
        random.shuffle(available)
        chosen = available[:n_boards]
        for p in chosen:
            participant_usage[p] += 1
        plan_data.append((day, slot, sc_name, chosen))
    return plan_data

def create_plan():
    boards = st.session_state.boards
    participants = st.session_state.participants
    scenarios = st.session_state.scenarios
    days_of_week = st.session_state.days_of_week
    timeslots = st.session_state.timeslots

    if not boards:
        st.error("En az bir sektör (Board) olmalı.")
        return
    if not participants:
        st.error("En az bir katılımcı olmalı.")
        return
    if not scenarios:
        st.error("En az bir senaryo satırı olmalı.")
        return

    # 1) Tekrarlı senaryoları genişlet
    expanded_scenarios = []
    for (sc_name, sc_day, sc_slot, sc_rep) in scenarios:
        for _ in range(sc_rep):
            expanded_scenarios.append((sc_day, sc_slot, sc_name))

    # 2) Sıralama
    def day_key(day_str):
        return days_of_week.index(day_str) if day_str in days_of_week else 9999
    def slot_key(slot_str):
        return timeslots.index(slot_str) if slot_str in timeslots else 9999

    expanded_scenarios.sort(key=lambda x: (day_key(x[0]), slot_key(x[1]), x[2]))

    # 3) Atama
    method = st.session_state.selected_method
    if method == "Random":
        plan_data = assign_random(expanded_scenarios, participants, boards)
    elif method == "Round Robin":
        plan_data = assign_round_robin(expanded_scenarios, participants, boards)
    elif method == "Balanced":
        plan_data = assign_balanced(expanded_scenarios, participants, boards)
    elif method == "Constraint (Latin Square)":
        plan_data = assign_constraint_latin(expanded_scenarios, participants, boards)
    else:
        plan_data = assign_random(expanded_scenarios, participants, boards)

    if plan_data is not None:
        st.session_state.plan_data = plan_data
        st.success(f"Plan oluşturuldu! (Yöntem: {method})")

# -------------------------------------------------------------------------
# 9) Excel'e Aktarma
# -------------------------------------------------------------------------
def export_to_excel():
    """Plan verilerini Excel olarak indirilebilecek hale getir."""
    plan_data = st.session_state.plan_data
    boards = st.session_state.boards
    participants = st.session_state.participants

    if not plan_data:
        st.error("Önce 'Plan Oluştur' butonuna basın.")
        return

    wb = openpyxl.Workbook()
    ws_plan = wb.active
    ws_plan.title = "Roster Plan"

    # Başlıklar
    ws_plan.cell(row=1, column=1).value = "Gün"
    ws_plan.cell(row=1, column=2).value = "Zaman"
    ws_plan.cell(row=1, column=3).value = "Senaryo"
    for i, b_name in enumerate(boards, start=4):
        ws_plan.cell(row=1, column=i).value = b_name

    color_palette = [
        "FFB6C1", "87CEFA", "98FB98", "FFA07A", "DDA0DD", "F0E68C",
        "FFA500", "B0E0E6", "FFD700", "90EE90", "FF69B4", "6495ED"
    ]
    participant_color = {}
    for i, p in enumerate(participants):
        participant_color[p] = color_palette[i % len(color_palette)]

    row_idx = 2
    for (day, slot, sc_name, assigned) in plan_data:
        ws_plan.cell(row=row_idx, column=1).value = day
        ws_plan.cell(row=row_idx, column=2).value = slot
        ws_plan.cell(row=row_idx, column=3).value = sc_name
        for col_idx, p in enumerate(assigned, start=4):
            ws_plan.cell(row=row_idx, column=col_idx).value = p
            fill_color = participant_color.get(p, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws_plan.cell(row=row_idx, column=col_idx).fill = fill
        row_idx += 1

    # Summary sayfası
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1).value = "Katılımcı"
    for i, b_name in enumerate(boards, start=2):
        ws_summary.cell(row=1, column=i).value = b_name
    last_col = len(boards) + 2
    ws_summary.cell(row=1, column=last_col).value = "Toplam"

    participant_board_counts = {p: {b: 0 for b in boards} for p in participants}
    for (_, _, _, assigned) in plan_data:
        for b_idx, p in enumerate(assigned):
            b_name = boards[b_idx]
            participant_board_counts[p][b_name] += 1

    row_idx = 2
    for p in participants:
        ws_summary.cell(row=row_idx, column=1).value = p
        fill_color = participant_color.get(p, "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws_summary.cell(row=row_idx, column=1).fill = fill

        total_count = 0
        for col_idx, b in enumerate(boards, start=2):
            c = participant_board_counts[p][b]
            ws_summary.cell(row=row_idx, column=col_idx).value = c
            total_count += c
        ws_summary.cell(row=row_idx, column=last_col).value = total_count
        row_idx += 1

    # Byte olarak kaydet ve download_button ile sun
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    st.download_button(
        label="Excel olarak indir",
        data=excel_data,
        file_name="roster_plan.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -------------------------------------------------------------------------
# 10) Ana Uygulama (main)
# -------------------------------------------------------------------------
def main():
    st.set_page_config(layout="wide")
    st.title("Roster Planlama")

    initialize_session_states()

    # Üst kısım: katılımcılar, sektörler, günler, zaman aralıkları, standart senaryolar
    col_part, col_board, col_day, col_time, col_std_s = st.columns(5)

    # ------------------ Katılımcılar ------------------
    with col_part:
        st.subheader("Katılımcılar")
        participants = st.session_state.participants
        st.write(participants)

        if st.button("Katılımcı Ekle", key="add_participant_btn"):
            add_participant()

        remove_participant_selected = st.selectbox(
            "Silinecek Katılımcı",
            options=[""] + participants,
            key="remove_select_participant"
        )
        if st.button("Katılımcıyı Sil", key="remove_participant_btn"):
            if remove_participant_selected:
                remove_participant(remove_participant_selected)

    # ------------------ Sektörler (Boards) ------------------
    with col_board:
        st.subheader("Sektörler (Boards)")
        boards = st.session_state.boards
        st.write(boards)

        # Ekleme
        add_board_col1, add_board_col2 = st.columns([3, 1])
        with add_board_col1:
            new_board_name = st.text_input("Yeni Sektör Adı", key="new_board_name")
        with add_board_col2:
            if st.button("Ekle", key="add_board_btn"):
                add_board(new_board_name)

        # Düzenleme
        if boards:
            st.markdown("### Sektör Güncelle")
            edit_board_selected = st.selectbox("Düzenlenecek Sektör", options=boards, key="edit_select_board")
            edit_new_board_name = st.text_input("Yeni Sektör Adı", key="edit_board_name")
            if st.button("Güncelle", key="edit_board_btn"):
                if edit_new_board_name:
                    edit_board(edit_board_selected, edit_new_board_name)

            # Silme
            st.markdown("### Sektör Sil")
            remove_board_selected = st.selectbox("Silinecek Sektör", options=boards, key="remove_select_board")
            if st.button("Sektörü Sil", key="remove_board_btn"):
                remove_board(remove_board_selected)

    # ------------------ Günler ------------------
    with col_day:
        st.subheader("Günler")
        days = st.session_state.days_of_week
        st.write(days)

        # Ekleme
        add_day_col1, add_day_col2 = st.columns([3, 1])
        with add_day_col1:
            new_day_name = st.text_input("Yeni Gün Adı", key="new_day")
        with add_day_col2:
            if st.button("Ekle", key="add_day_btn"):
                add_day(new_day_name)

        # Düzenleme
        if days:
            st.markdown("### Gün Güncelle")
            edit_day_selected = st.selectbox("Düzenlenecek Gün", options=days, key="edit_select_day")
            edit_new_day_val = st.text_input("Yeni Gün Adı", key="edit_day_name")
            if st.button("Güncelle", key="edit_day_btn"):
                if edit_new_day_val:
                    rename_day(edit_day_selected, edit_new_day_val)

            # Silme
            st.markdown("### Gün Sil")
            remove_day_selected = st.selectbox("Silinecek Gün", options=days, key="remove_select_day")
            if st.button("Günü Sil", key="remove_day_btn"):
                remove_day(remove_day_selected)

    # ------------------ Zaman Aralıkları (Timeslots) ------------------
    with col_time:
        st.subheader("Zaman Aralıkları")
        timeslots = st.session_state.timeslots
        st.write(timeslots)

        # Ekleme
        add_timeslot_col1, add_timeslot_col2 = st.columns([3, 1])
        with add_timeslot_col1:
            new_timeslot_val = st.text_input("Yeni Zaman Aralığı", key="new_timeslot")
        with add_timeslot_col2:
            if st.button("Ekle", key="add_timeslot_btn"):
                add_timeslot(new_timeslot_val)

        # Düzenleme
        if timeslots:
            st.markdown("### Zaman Aralığı Güncelle")
            edit_timeslot_selected = st.selectbox("Düzenlenecek Zaman Aralığı", options=timeslots, key="edit_select_timeslot")
            edit_new_timeslot_val = st.text_input("Yeni Zaman Aralığı", key="edit_timeslot_name")
            if st.button("Güncelle", key="edit_timeslot_btn"):
                if edit_new_timeslot_val:
                    edit_timeslot(edit_timeslot_selected, edit_new_timeslot_val)

            # Silme
            st.markdown("### Zaman Aralığı Sil")
            remove_timeslot_selected = st.selectbox("Silinecek Zaman Aralığı", options=timeslots, key="remove_select_timeslot")
            if st.button("Zaman Aralığını Sil", key="remove_timeslot_btn"):
                remove_timeslot(remove_timeslot_selected)

    # ------------------ Standart Senaryolar ------------------
    with col_std_s:
        st.subheader("Standart Senaryolar")
        std_scenarios = st.session_state.standard_scenarios
        st.write(std_scenarios)

        # Ekleme
        add_std_scen_col1, add_std_scen_col2 = st.columns([3, 1])
        with add_std_scen_col1:
            new_std_scenario = st.text_input("Yeni Standart Senaryo Adı", key="new_std_scenario")
        with add_std_scen_col2:
            if st.button("Ekle", key="add_std_scenario_btn"):
                add_standard_scenario(new_std_scenario)

        # Düzenleme
        if std_scenarios:
            st.markdown("### Standart Senaryo Güncelle")
            edit_std_scen_selected = st.selectbox("Düzenlenecek Standart Senaryo", options=std_scenarios, key="edit_select_std_scenario")
            edit_new_std_scen_val = st.text_input("Yeni Standart Senaryo Adı", key="edit_std_scenario_name")
            if st.button("Güncelle", key="edit_std_scenario_btn"):
                if edit_new_std_scen_val:
                    edit_standard_scenario(edit_std_scen_selected, edit_new_std_scen_val)

            # Silme
            st.markdown("### Standart Senaryo Sil")
            remove_std_scen_selected = st.selectbox("Silinecek Standart Senaryo", options=std_scenarios, key="remove_select_std_scenario")
            if st.button("Standart Senaryoyu Sil", key="remove_std_scenario_btn"):
                remove_standard_scenario(remove_std_scen_selected)

    # ------------------ Asıl Planlanacak Senaryolar ------------------
    st.write("---")
    st.subheader("Asıl Planlanacak Senaryolar")
    scenarios_copy = st.session_state.scenarios.copy()
    for idx, scenario in enumerate(scenarios_copy):
        col_sc1, col_sc2, col_sc3, col_sc4, col_sc5 = st.columns([2,1,1,1,0.5])
        sc_name, sc_day, sc_slot, sc_rep = scenario

        with col_sc1:
            new_sc_name = st.selectbox(
                f"Senaryo Adı [{idx}]",
                options=st.session_state.standard_scenarios or ["Senaryo?"],
                index=(st.session_state.standard_scenarios.index(sc_name)
                       if sc_name in st.session_state.standard_scenarios else 0),
                key=f"sc_name_{idx}"
            )
        with col_sc2:
            new_day = st.selectbox(
                f"Gün [{idx}]",
                options=st.session_state.days_of_week or ["Gün?"],
                index=(st.session_state.days_of_week.index(sc_day)
                       if sc_day in st.session_state.days_of_week else 0),
                key=f"sc_day_{idx}"
            )
        with col_sc3:
            new_slot = st.selectbox(
                f"Zaman [{idx}]",
                options=st.session_state.timeslots or ["Zaman?"],
                index=(st.session_state.timeslots.index(sc_slot)
                       if sc_slot in st.session_state.timeslots else 0),
                key=f"sc_slot_{idx}"
            )
        with col_sc4:
            new_rep = st.number_input(
                f"Tekrar [{idx}]",
                min_value=1, max_value=20, value=sc_rep, step=1,
                key=f"sc_rep_{idx}"
            )
        with col_sc5:
            if st.button(f"Sil [{idx}]", key=f"remove_scenario_{idx}"):
                remove_scenario(idx)

        # Listeyi güncelle
        if idx < len(st.session_state.scenarios):
            st.session_state.scenarios[idx] = [new_sc_name, new_day, new_slot, new_rep]

    if st.button("Yeni Senaryo Satırı Ekle", key="add_new_scenario_btn"):
        add_scenario()

    # ------------------ Atama Yöntemi Seçimi & Plan Oluşturma ------------------
    st.write("---")
    st.subheader("Atama Yöntemi Seçimi")
    st.session_state.selected_method = st.selectbox(
        "Yöntem",
        st.session_state.assignment_methods,
        index=st.session_state.assignment_methods.index(st.session_state.selected_method)
        if st.session_state.selected_method in st.session_state.assignment_methods else 0,
        key="assignment_method_selectbox"
    )

    if st.button("Plan Oluştur", key="create_plan_btn"):
        create_plan()

    # Eğer plan oluşturulmuşsa, tabloyu göster
    if st.session_state.plan_data:
        st.write("**Oluşturulan Plan**")
        df_data = []
        columns = ["Gün", "Zaman", "Senaryo"]
        columns.extend(st.session_state.boards)

        for (day, slot, sc_name, assigned) in st.session_state.plan_data:
            df_data.append([day, slot, sc_name] + assigned)

        df = pd.DataFrame(df_data, columns=columns)
        st.dataframe(df)

        st.write("Planı Excel formatında indirmek için butona tıklayabilirsiniz:")
        export_to_excel()

# -------------------------------------------------------------------------
# 11) Çalıştırma
# -------------------------------------------------------------------------
if __name__ == "__main__":
    main()